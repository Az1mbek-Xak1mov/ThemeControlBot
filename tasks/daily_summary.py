import os
import sys
BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, BASE)
os.chdir(BASE)
from dotenv import load_dotenv
load_dotenv(os.path.join(BASE, ".env"))
from environment.utils import Env
import asyncio
from aiogram import Bot
from aiogram.types import FSInputFile
from openpyxl import Workbook
admin_chat_id = Env().bot.ADMIN_CHAT_ID
TOKEN = Env().bot.SUMMARY_TOKEN
bot = Bot(TOKEN)

from db.manager import total_messages, get_messages_for_chat, get_all_groups, get_user_objects_for_group

async def create_activity_report(filename="daily_report.xlsx"):
    wb = Workbook()
    groups = await get_all_groups()

    helper = []
    for idx, group in enumerate(groups):
        if idx == 0:
            ws = wb.active
        else:
            ws = wb.create_sheet()
        ws.title = group.title
        ws.append(["🙋 Ism", "👤 Username", "🔢 Sanoq", "💯 Foiz"])
        info_user = await get_user_objects_for_group(group.chat_id)
        total = await total_messages(1, group.chat_id)
        for info in info_user:
            count = await get_messages_for_chat(1, group.chat_id, info.chat_id, )
            if total!=0:
                percent = f"{(100 / total * count):.2f}%"
            else:
                percent=0
            if not info.username:
                info.username = '-'
            helper.append([info.name, info.username, count, percent])
            ws.append([info.name, info.username, count, percent])
        sorted_data = sorted(helper, key=lambda r: r[2], reverse=True)

        top5 = sorted_data[:5]
        bottom5 = sorted_data[-5:] if len(sorted_data) >= 5 else sorted_data

        start_col = 7

        ws.cell(row=1, column=start_col, value="🏆 Top 5")
        ws.cell(row=1, column=start_col + 1, value="🔢 Sanoq")
        for idx, row in enumerate(top5, start=2):
            ws.cell(row=idx, column=start_col, value=row[0])
            ws.cell(row=idx, column=start_col + 1, value=row[2])

        offset = len(top5) + 3
        ws.cell(row=offset, column=start_col, value="📉 Bottom 5")
        ws.cell(row=offset, column=start_col + 1, value="🔢 Sanoq")
        for idx, row in enumerate(bottom5, start=offset + 1):
            ws.cell(row=idx, column=start_col, value=row[0])
            ws.cell(row=idx, column=start_col + 1, value=row[2])
        helper = []

        from openpyxl.utils import get_column_letter
        for col in ws.columns:
            max_len = max((len(str(c.value)) for c in col if c.value), default=2)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2

    wb.save(filename)
    return filename


async def send_summary_to_admin_day():
    filepath = await create_activity_report()
    file = FSInputFile(filepath)
    await bot.send_document(admin_chat_id, file, caption="📈 Kunlik hisobot")
    await bot.session.close()


if __name__ == "__main__":
    asyncio.run(send_summary_to_admin_day())
